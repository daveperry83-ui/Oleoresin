"""Ingesta de ``Product Reference Internal.xlsx`` (first choice list).

Además de la hoja principal de oleorresinas, este archivo trae la tabla de
equivalencias con productos de la competencia (Kalsec / Mane), que es el
insumo del buscador de reemplazo.
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional

import openpyxl

from core.units import normalize_solubility, normalize_yesno, parse_spec_value
from data_layer.schema import SOURCE_FIRST_CHOICE, STATUS_ACTIVE, Product

MAIN_SHEET = "standard oleoresin "
ORGANIC_SHEET = "Organic products"
COMPETITOR_SHEET = "Kalsec reference"

#: Palabras del nombre comercial que identifican la familia botánica.
FAMILY_HINTS = [
    "allspice", "pimento", "anise", "annatto", "basil", "bay", "capsicum",
    "caraway", "cardamom", "cassia", "celery", "chipotle", "cilantro",
    "cinnamon", "clove", "coriander", "cumin", "dill", "fennel", "foenugreek",
    "garlic", "ginger", "habanero", "jalapeno", "juniper", "lovage", "mace",
    "marjoram", "mustard", "nutmeg", "onion", "oregano", "paprika", "parsley",
    "pepper", "rosemary", "sage", "tarragon", "thyme", "turmeric", "vanilla",
    "wasabi", "beta carotene",
]


def _clean(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def family_from_description(description: str) -> str:
    text = description.lower()
    if "white pepper" in text:
        return "pepper, white"
    if "black pepper" in text:
        return "pepper, black and green"
    for hint in FAMILY_HINTS:
        if hint in text:
            return hint
    return ""


def load(path: str | Path) -> List[Product]:
    wb = openpyxl.load_workbook(path, data_only=True)
    products: List[Product] = []

    ws = wb[MAIN_SHEET]
    for row in ws.iter_rows(min_row=2):
        code = _clean(row[0].value)
        if not code:
            continue
        description = _clean(row[1].value) or code
        analytes = {}
        for idx, analyte in ((5, "volatile_oil"), (6, "capsaicin"), (7, "colour"), (8, "piperine")):
            if idx < len(row):
                rng = parse_spec_value(row[idx].value, analyte=analyte)
                if rng is not None:
                    analytes[analyte] = rng

        # La columna "Colour" de cúrcuma en realidad declara curcumina.
        if "turmeric" in description.lower() and "colour" in analytes:
            raw = analytes["colour"].raw.lower()
            if "curcumin" in raw:
                analytes["curcumin"] = analytes.pop("colour")

        contains_gmo = normalize_yesno(row[2].value if len(row) > 2 else None)
        products.append(
            Product(
                code=code,
                description=description,
                source=SOURCE_FIRST_CHOICE,
                family=family_from_description(description),
                sheet=MAIN_SHEET.strip(),
                solubility=normalize_solubility(row[3].value if len(row) > 3 else None),
                gmo_free=(not contains_gmo) if contains_gmo is not None else None,
                allergen_free=normalize_yesno(row[4].value if len(row) > 4 else None),
                status=STATUS_ACTIVE,
                notes=_clean(row[10].value) if len(row) > 10 else "",
                analytes=analytes,
            )
        )

    if ORGANIC_SHEET in wb.sheetnames:
        wso = wb[ORGANIC_SHEET]
        for row in wso.iter_rows(min_row=2, values_only=True):
            code = _clean(row[0])
            if not code or code.lower() == "code":
                continue
            description = _clean(row[1]) or code
            analytes = {}
            lo = parse_spec_value(row[4] if len(row) > 4 else None, analyte="volatile_oil")
            hi = parse_spec_value(row[5] if len(row) > 5 else None, analyte="volatile_oil")
            if lo and hi:
                from core.units import SpecRange

                analytes["volatile_oil"] = SpecRange(
                    lo.midpoint, hi.midpoint, "range", "%", f"{lo.raw} – {hi.raw}"
                )
            elif lo:
                analytes["volatile_oil"] = lo
            products.append(
                Product(
                    code=code,
                    description=description,
                    source=SOURCE_FIRST_CHOICE,
                    family=family_from_description(description),
                    sheet=ORGANIC_SHEET,
                    solubility=normalize_solubility(row[3] if len(row) > 3 else None),
                    organic=True,
                    status=STATUS_ACTIVE,
                    analytes=analytes,
                )
            )

    return products


@dataclass
class CompetitorMatch:
    competitor: str
    competitor_code: str
    competitor_desc: str
    robertet_code: str
    verdict: str

    @property
    def has_offer(self) -> bool:
        return bool(self.robertet_code)

    @property
    def confidence(self) -> str:
        v = self.verdict.lower()
        if not self.robertet_code:
            return "none"
        if "1:1" in v or "match the spec" in v or "within the spec" in v:
            return "high"
        if "closest" in v or "approved" in v or "in range" in v:
            return "medium"
        return "medium"


def load_competitor_map(path: str | Path) -> List[CompetitorMatch]:
    """Lee la hoja ``Kalsec reference``: dos bloques, Kalsec y Mane."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if COMPETITOR_SHEET not in wb.sheetnames:
        return []
    ws = wb[COMPETITOR_SHEET]
    out: List[CompetitorMatch] = []
    blocks = [("Kalsec", 0, 1, 2, 3), ("Mane", 6, 7, 8, 9)]
    for row in ws.iter_rows(min_row=2, values_only=True):
        for name, i_desc, i_code, i_rob, i_note in blocks:
            if len(row) <= i_note:
                continue
            desc, code = _clean(row[i_desc]), _clean(row[i_code])
            if not desc and not code:
                continue
            out.append(
                CompetitorMatch(
                    competitor=name,
                    competitor_code=code,
                    competitor_desc=desc,
                    robertet_code=_clean(row[i_rob]),
                    verdict=_clean(row[i_note]),
                )
            )
    return out
