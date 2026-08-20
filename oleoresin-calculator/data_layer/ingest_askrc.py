"""Ingesta del ASKRC (extended list).

Dos trampas del archivo que este módulo resuelve explícitamente:

1. **El formato es dato.** La hoja ``Legend`` documenta que el tachado significa
   "Void / Do Not Sample". ``pandas.read_excel`` descarta esa información en
   silencio, así que un parser convencional recomendaría producto
   descontinuado con total confianza. Aquí se lee ``cell.font.strike``.

2. **El encabezado se mueve.** La fila con ``Code`` en la columna A varía entre
   la 3 y la 10 según la hoja. Se detecta, no se asume.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl

from core.units import normalize_solubility, normalize_yesno, parse_spec_value
from data_layer.schema import (
    SOURCE_EXTENDED,
    STATUS_ACTIVE,
    STATUS_CONVERTED,
    STATUS_VOID,
    Product,
)

#: Hojas que no son catálogo de producto.
NON_PRODUCT_SHEETS = {
    "Sheet2",
    "Legend",
    "Kosher Passover",
    "Compatibility Report",
    "User Notes",
    "Robertet USA Products",
}

#: Los encabezados del ASKRC no son consistentes entre hojas: "VO Max" convive
#: con "MAX-VO", "Current Shelf Life (month)" con "(months)". Se normalizan
#: contra este mapa antes de leer nada.
COLUMN_ALIASES: Dict[str, str] = {
    "code": "code",
    "product code": "code",
    "converted to": "converted_to",
    "product description": "description",
    "description": "description",
    "description (standard)": "description",
    "water /oil dispersable": "solubility",
    "water/oil dispersable": "solubility",
    "water /oil dispersible": "solubility",
    "water/oil dispers-ability": "solubility",
    "water/oil dispers-ab": "solubility",
    "solubility": "solubility",
    "form": "form",
    "legal status": "legal_status",
    "kosher": "kosher",
    "halal": "halal",
    "vegan": "vegan",
    "gmo free": "gmo_free",
    "contains gmo ingredients": "contains_gmo",
    "allergens": "allergens",
    "allergen free (us, ca, eu)": "allergen_free",
    "contains soybean oil": "contains_soy",
    "**contains refined soybean oil": "contains_soy",
    "dosage level": "dosage_level",
    "current shelf life (month)": "shelf_life",
    "current shelf life (months)": "shelf_life",
    "note": "notes",
    "notes": "notes",
}

#: El ASKRC nombra la misma columna de doce maneras ("VO Max", "MAX-VO",
#: "%VO Max", "Max - Volatile Oil Content"…). Enumerar variantes no escala, así
#: que el analito y el extremo se deducen por patrón.
_ANALYTE_PATTERNS = (
    # El orden importa: "max-caps scoville unit" contiene ambas palabras.
    (r"scoville", "scoville", False),
    (r"carnos", "carnosic_acid", True),
    (r"curcumin", "curcumin", True),
    (r"piperine", "piperine", True),
    (r"\bcaps(aicin)?\b", "capsaicin", True),
    (r"colou?r", "colour", False),
    (r"vanillin", "vanillin", True),
    (r"(^|[^a-z])vo([^a-z]|$)|volatile\s*oil", "volatile_oil", True),
)

#: Encabezados que parecen analito pero no lo son.
_ANALYTE_EXCLUDE = re.compile(r"soluble solid|column\d|^unit$|^type$|dosage")


def classify_analyte_header(header: str):
    """('MAX-Caps Scoville Unit') -> ('scoville', 'high', False) | None.

    Devuelve (analito, extremo, es_fracción). ``extremo`` es ``high``, ``low``,
    ``calc`` o ``both``.
    """
    h = header.strip().lower()
    if not h or _ANALYTE_EXCLUDE.search(h):
        return None

    analyte = fraction = None
    for pattern, name, is_fraction in _ANALYTE_PATTERNS:
        if re.search(pattern, h):
            analyte, fraction = name, is_fraction
            break
    if analyte is None:
        return None

    if re.search(r"\bcalc(ulated)?\b|^calc", h):
        slot = "calc"
    elif re.search(r"\bmax\b|max\s*-|-\s*max|^max", h):
        slot = "high"
    elif re.search(r"\bmin\b|min\s*-|-\s*min|^min", h):
        slot = "low"
    else:
        slot = "both"
    return analyte, slot, fraction


def _clean(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _norm_header(value) -> str:
    return _clean(value).lower().replace("\n", " ").strip()


def family_from_sheet(sheet_name: str) -> str:
    """'31-Pepper, Black and Green' -> 'pepper, black and green'.

    Los nombres de hoja mezclan prefijos numéricos ('28-Paprika'), sufijos de
    familia ('43-47s Oleoresin Blends') y alias separados por barra
    ('00-Allspice | Pimento'). Se queda el primer alias, sin numeración.
    """
    name = str(sheet_name).strip()
    name = re.sub(r"^[\d\s\-–]+", "", name)      # '43-47s Oleoresin' -> 's Oleoresin'
    name = re.sub(r"^s\b\s*", "", name)          # sufijo de decena
    name = name.split("|")[0]                     # 'Allspice | Pimento' -> 'Allspice'
    name = re.sub(r"\s+", " ", name).strip(" -–").lower()
    return name


def find_header_row(ws, max_scan: int = 15) -> Optional[int]:
    """Localiza la fila del encabezado buscando 'Code' en la columna A."""
    for row in range(1, min(ws.max_row, max_scan) + 1):
        if _clean(ws.cell(row, 1).value).lower() == "code":
            return row
    return None


def _build_range(parts: dict, analyte: str, fraction: bool):
    """Reconstruye un SpecRange a partir de columnas Max/Min/Calc separadas."""
    low = parts.get("low")
    high = parts.get("high")
    calc = parts.get("calc")
    both = parts.get("both")

    lo = parse_spec_value(low, analyte=analyte, assume_fraction=fraction) if low is not None else None
    hi = parse_spec_value(high, analyte=analyte, assume_fraction=fraction) if high is not None else None

    if lo is not None and hi is not None:
        from core.units import SpecRange

        raw = f"{lo.raw} – {hi.raw}"
        return SpecRange(lo.midpoint, hi.midpoint, "range", lo.unit or hi.unit, raw)
    if lo is not None:
        from core.units import SpecRange

        return SpecRange(lo.midpoint, None, "min", lo.unit, lo.raw)
    if hi is not None:
        from core.units import SpecRange

        return SpecRange(None, hi.midpoint, "max", hi.unit, hi.raw)
    if both is not None:
        return parse_spec_value(both, analyte=analyte, assume_fraction=fraction)
    if calc is not None:
        return parse_spec_value(calc, analyte=analyte, assume_fraction=fraction)
    return None


def load(path: str | Path) -> List[Product]:
    """Lee el ASKRC completo conservando la semántica del formato."""
    wb = openpyxl.load_workbook(path, data_only=True)
    products: List[Product] = []

    for ws in wb.worksheets:
        if ws.title in NON_PRODUCT_SHEETS:
            continue
        header_row = find_header_row(ws)
        if header_row is None:
            continue

        family = family_from_sheet(ws.title)
        organic = "organic" in ws.title.lower()
        ncols = min(ws.max_column, 80)
        headers = {c: _norm_header(ws.cell(header_row, c).value) for c in range(1, ncols + 1)}

        for row in range(header_row + 1, ws.max_row + 1):
            code_cell = ws.cell(row, 1)
            code = _clean(code_cell.value)
            if not code or code.lower() == "code":
                continue

            fields: Dict[str, object] = {}
            analyte_parts: Dict[str, Dict[str, object]] = {}
            analyte_fraction: Dict[str, bool] = {}

            for col, header in headers.items():
                if not header:
                    continue
                value = ws.cell(row, col).value
                if header in COLUMN_ALIASES:
                    fields.setdefault(COLUMN_ALIASES[header], value)
                    continue
                hit = classify_analyte_header(header)
                if hit is not None:
                    analyte, slot, fraction = hit
                    if value is not None and _clean(value):
                        analyte_parts.setdefault(analyte, {})[slot] = value
                        analyte_fraction[analyte] = fraction
                elif re.search(r"dispers|solub", header):
                    fields.setdefault("solubility", value)

            analytes = {}
            for analyte, parts in analyte_parts.items():
                rng = _build_range(parts, analyte, analyte_fraction.get(analyte, False))
                if rng is not None:
                    analytes[analyte] = rng

            # El ASKRC declara Scoville solo en algunas hojas de capsicum; donde
            # falta se deriva de la capsaicina con el mismo factor del archivo.
            if "capsaicin" in analytes and "scoville" not in analytes:
                from core.units import SpecRange, capsaicin_to_shu

                caps = analytes["capsaicin"]
                analytes["scoville"] = SpecRange(
                    capsaicin_to_shu(caps.low) if caps.low is not None else None,
                    capsaicin_to_shu(caps.high) if caps.high is not None else None,
                    caps.kind,
                    "SHU",
                    f"derivado de capsaicina {caps.format()}",
                )

            # --- el formato como dato -------------------------------------
            struck = bool(code_cell.font and code_cell.font.strike)
            converted_to = _clean(fields.get("converted_to"))
            if struck:
                status = STATUS_VOID
            elif converted_to:
                status = STATUS_CONVERTED
            else:
                status = STATUS_ACTIVE

            gmo_free = normalize_yesno(fields.get("gmo_free"))
            if gmo_free is None:
                contains_gmo = normalize_yesno(fields.get("contains_gmo"))
                gmo_free = (not contains_gmo) if contains_gmo is not None else None

            allergens = normalize_yesno(fields.get("allergens"))
            allergen_free = normalize_yesno(fields.get("allergen_free"))
            if allergen_free is None and allergens is not None:
                allergen_free = not allergens

            products.append(
                Product(
                    code=code,
                    description=_clean(fields.get("description")) or code,
                    source=SOURCE_EXTENDED,
                    family=family,
                    sheet=ws.title,
                    solubility=normalize_solubility(fields.get("solubility")),
                    form=_clean(fields.get("form")),
                    legal_status=_clean(fields.get("legal_status")),
                    kosher=normalize_yesno(fields.get("kosher")),
                    halal=normalize_yesno(fields.get("halal")),
                    vegan=normalize_yesno(fields.get("vegan")),
                    gmo_free=gmo_free,
                    allergen_free=allergen_free,
                    contains_soy=normalize_yesno(fields.get("contains_soy")),
                    organic=organic,
                    shelf_life_months=_shelf_life(fields.get("shelf_life")),
                    dosage_level=_clean(fields.get("dosage_level")),
                    status=status,
                    converted_to=converted_to,
                    notes=_clean(fields.get("notes")),
                    analytes=analytes,
                )
            )

    return products


def _shelf_life(value) -> Optional[float]:
    rng = parse_spec_value(value, analyte=None)
    return rng.midpoint if rng else None
